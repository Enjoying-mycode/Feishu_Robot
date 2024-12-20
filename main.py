"""通过配置文件在指定飞书群发送报表截图、点评、文件"""

import basic_finction
# openpyxl是一个纯 Python 库，不依赖于任何外部的 COM 组件或 Office 安装，直接通过API来操作Excel文件
from openpyxl import load_workbook
import os
import get_screenshot_from_excel
import get_text_from_excel
import time
import datetime


def ensure_dir_exists(dir_path):
    # 检查目录是否存在
    if not os.path.exists(dir_path):
        # 目录不存在，创建目录
        os.makedirs(dir_path)
    else:
        pass


while True:
    # 记录开始时间
    start_time = datetime.datetime.now()
    cgf_file_path = r'D:\MyDocuments\wangjia93\桌面\测试文件\config.xlsx'
    workbook = load_workbook(filename=cgf_file_path, data_only=True)
    sheet = workbook.active  # 假设我们读取的是活动工作表
    max_row = sheet.max_row     # 最大行数

    num = range(2, max_row + 1)
    for row in num:
        send_type = sheet["A"+str(row)].value
        # 获取群名
        chat_name = sheet["B"+str(row)].value
        # 获取文件导出路径
        excel_file_path = sheet["C"+str(row)].value
        # 获取excel文件名
        excel_file_name = sheet["D"+str(row)].value
        # 获取截图存放文件夹名
        image_file_path = sheet["E" + str(row)].value

        entries = os.listdir(excel_file_path)
        # 判断cfg文件中的导出文件是否在
        if excel_file_name in entries:
            if send_type == '报表截图':
                # excel文件绝对路径
                excel_path = os.path.join(excel_file_path, excel_file_name)
                # 截图名称
                image_name = '截图-' + sheet["D"+str(row)].value.replace('.xlsx','') + '.png'
                # 截图文件导出文件夹
                image_path = os.path.join(image_file_path, str(chat_name))
                # 截图存放路径是否存在，不存在，则新建
                ensure_dir_exists(image_path)
                # 截图文件绝对路径
                output_image_path = os.path.join(image_path, image_name)

                # 捕获Excel文件的报表截图
                get_screenshot_from_excel.capture_excel_data_as_image(excel_path, output_image_path)

                # 发送截图至飞书群
                basic_finction.send_image(image_path, image_name, chat_name)

                # 删除截图
                os.remove(output_image_path)

                # 删除源文件
                os.remove(excel_path)
            elif send_type == '评论':
                # excel文件绝对路径
                excel_path = os.path.join(excel_file_path, excel_file_name)

                # 依次获取A列单元格文本并拼接成文本，并去掉字符串中多余的符号
                excel_text = get_text_from_excel.capture_excel_cola_data_as_text(excel_path).replace("&", "")

                # 发送文本消息至飞书群
                basic_finction.send_text(excel_text, chat_name)
                # 删除源文件
                os.remove(excel_path)
            elif send_type == '文件':
                # excel文件绝对路径
                excel_path = os.path.join(excel_file_path, excel_file_name)
                # 发送文件
                basic_finction.send_file(excel_file_path, excel_file_name, chat_name)
                # 等待10秒，防止文件过大，导致发送时间长
                # time.sleep(5)
                # 删除源文件
                os.remove(excel_path)
        else:
            pass

    workbook.close()

    # 记录结束时间
    end_time = datetime.datetime.now()

    # 总运行时间
    cost_time = (end_time - start_time).total_seconds()

    print(f'开始时间：{start_time}，结束时间：{end_time}，总花费时间：{cost_time}')

    print('20秒后重启')
    time.sleep(20)








